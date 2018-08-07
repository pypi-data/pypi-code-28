import csv
import os
import subprocess

from django.core.management.base import BaseCommand
from election.models import Race
from geography.models import Division, DivisionLevel
from openpyxl import load_workbook
from time import sleep

from raceratings.models import Author, Category, DataProfile, RaceRating


class Command(BaseCommand):
    help = (
        'Bootstrap development.'
    )

    def create_categories(self):
        Category.objects.get_or_create(
            label='Solid Republican',
            short_label='Solid-R'
        )

        Category.objects.get_or_create(
            label='Likely Republican',
            short_label='Likely-R'
        )

        Category.objects.get_or_create(
            label='Lean Republican',
            short_label='Lean-R'
        )

        Category.objects.get_or_create(
            label='Toss-Up',
            short_label='Toss-Up'
        )

        Category.objects.get_or_create(
            label='Lean Democrat',
            short_label='Lean-D'
        )

        Category.objects.get_or_create(
            label='Likely Democrat',
            short_label='Likely-D'
        )

        Category.objects.get_or_create(
            label='Solid Democrat',
            short_label='Solid-D'
        )

        self.steve, created = Author.objects.get_or_create(
            first_name='Steve',
            last_name='Shepard'
        )

    def write_initial_csvs(self):
        wb_path = os.path.join(self.cmd_path, '../../bin/ratings.xlsx')
        wb = load_workbook(wb_path)

        for name in wb.sheetnames:
            csv_path = os.path.join(
                self.cmd_path, '../../bin/csv', '{}.csv'.format(name)
            )

            with open(csv_path, 'w') as writefile:
                subprocess.Popen(
                    [
                        'in2csv',
                        wb_path,
                        '-f',
                        'xlsx',
                        '--sheet',
                        name
                    ],
                    stdout=writefile
                )

    def get_initial_ratings(self):
        sleep(2)
        with open(os.path.join(self.cmd_path, '../../bin/csv/house.csv')) as f:
            reader = csv.DictReader(f)

            for row in reader:
                district_int = int(row['district'])
                if district_int < 10:
                    district_str = row['district'].zfill(2)
                else:
                    district_str = row['district']

                division = Division.objects.get(
                    level__name=DivisionLevel.DISTRICT,
                    code=district_str,
                    parent__label=row['state']
                )

                race = Race.objects.get(
                    office__division=division,
                    cycle__slug='2018',
                    special=False
                )

                data = {
                    'incumbent_party': row['party_hold']
                }

                data_profile, created = DataProfile.objects.get_or_create(
                    race=race,
                    data=data
                )

                incumbent_category = Category.objects.get(
                    short_label=row['2016 rating']
                )

                initial_category = Category.objects.get(
                    short_label=row['2018 rating']
                )

                incumbent, created = RaceRating.objects.get_or_create(
                    race=race,
                    author=self.steve,
                    category=incumbent_category,
                    incumbent=True
                )

                initial, created = RaceRating.objects.get_or_create(
                    race=race,
                    author=self.steve,
                    category=initial_category,
                    incumbent=False
                )

        with open(os.path.join(
            self.cmd_path, '../../bin/csv/senate.csv')
        ) as f:
            reader = csv.DictReader(f)

            for row in reader:
                if 'Special' in row['state']:
                    state = row['state'].split(' Special')[0]
                    division = Division.objects.get(
                        level__name=DivisionLevel.STATE,
                        label=state
                    )

                    race = Race.objects.get(
                        office__division=division,
                        office__body__slug='senate',
                        cycle__slug='2018',
                        special=True
                    )
                else:
                    division = Division.objects.get(
                        level__name=DivisionLevel.STATE,
                        label=row['state']
                    )
                    race = Race.objects.get(
                        office__division=division,
                        office__body__slug='senate',
                        cycle__slug='2018',
                        special=False
                    )

                data = {
                    'incumbent_party': row['party_hold']
                }

                data_profile, created = DataProfile.objects.get_or_create(
                    race=race,
                    data=data
                )

                incumbent_category = Category.objects.get(
                    short_label=row['2016 rating']
                )

                initial_category = Category.objects.get(
                    short_label=row['2018 rating']
                )

                incumbent, created = RaceRating.objects.get_or_create(
                    race=race,
                    author=self.steve,
                    category=incumbent_category,
                    incumbent=True
                )

                initial, created = RaceRating.objects.get_or_create(
                    race=race,
                    author=self.steve,
                    category=initial_category,
                    incumbent=False
                )

        with open(os.path.join(
            self.cmd_path, '../../bin/csv/governor.csv')
        ) as f:
            reader = csv.DictReader(f)

            for row in reader:
                division = Division.objects.get(
                    level__name=DivisionLevel.STATE,
                    label=row['state']
                )
                race = Race.objects.get(
                    office__division=division,
                    office__body=None,
                    cycle__slug='2018'
                )

                data = {
                    'incumbent_party': row['party_hold']
                }

                data_profile, created = DataProfile.objects.get_or_create(
                    race=race,
                    data=data
                )

                incumbent_category = Category.objects.get(
                    short_label=row['2016 rating']
                )

                initial_category = Category.objects.get(
                    short_label=row['2018 rating']
                )

                incumbent, created = RaceRating.objects.get_or_create(
                    race=race,
                    author=self.steve,
                    category=incumbent_category,
                    incumbent=True
                )

                initial, created = RaceRating.objects.get_or_create(
                    race=race,
                    author=self.steve,
                    category=initial_category,
                    incumbent=False
                )

    def handle(self, *args, **options):
        self.cmd_path = os.path.dirname(os.path.realpath(__file__))

        self.create_categories()
        self.write_initial_csvs()
        self.get_initial_ratings()
